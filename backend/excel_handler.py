import os
import shutil
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"G:\My Drive\Gonen\משק הראל\חשבון חשמל\חשבון חשמל.xlsx"
BACKUP_PATH = r"G:\My Drive\Gonen\משק הראל\חשבון חשמל\חשבון חשמל_backup.xlsx"

class ExcelHandler:
    def __init__(self, file_path=EXCEL_PATH):
        self.file_path = file_path

    def _create_backup(self):
        if os.path.exists(self.file_path):
            shutil.copy2(self.file_path, BACKUP_PATH)

    def get_history(self):
        """Reads all historical entries from the Excel spreadsheet for both meters."""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found at {self.file_path}")

        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        sheet = wb.active

        # Extract current rates in G1 and G2
        base_rate = sheet["G2"].value
        rate_with_vat = sheet["G1"].value

        # Read Big Watch (Main Meter / Tenant 1) rows
        big_watch_rows = []
        for r in range(4, sheet.max_row + 1):
            date_val = sheet.cell(row=r, column=3).value  # Col C
            reading_val = sheet.cell(row=r, column=4).value  # Col D
            if date_val is not None or reading_val is not None:
                big_watch_rows.append({
                    "row": r,
                    "no": sheet.cell(row=r, column=2).value,
                    "date": str(date_val).split(" ")[0] if date_val else "",
                    "reading": float(reading_val) if reading_val is not None else 0.0,
                    "kwh_month": sheet.cell(row=r, column=5).value,
                    "days": sheet.cell(row=r, column=6).value,
                    "kwh_day": sheet.cell(row=r, column=7).value,
                    "diff_kwh": sheet.cell(row=r, column=8).value,
                    "cost_col_i": sheet.cell(row=r, column=9).value  # Column I
                })

        # Read Small Watch (Sub-Meter / Tenant 2) rows
        small_watch_rows = []
        for r in range(4, sheet.max_row + 1):
            date_val = sheet.cell(row=r, column=12).value  # Col L
            reading_val = sheet.cell(row=r, column=13).value  # Col M
            if date_val is not None or reading_val is not None:
                small_watch_rows.append({
                    "row": r,
                    "no": sheet.cell(row=r, column=11).value,
                    "date": str(date_val).split(" ")[0] if date_val else "",
                    "reading": float(reading_val) if reading_val is not None else 0.0,
                    "kwh_month": sheet.cell(row=r, column=14).value,
                    "days": sheet.cell(row=r, column=15).value,
                    "kwh_day": sheet.cell(row=r, column=16).value,
                    "cost_col_q": sheet.cell(row=r, column=17).value  # Column Q
                })

        # Read historical tariffs table if present
        tariff_history = []
        if "תעריפי חשמל היסטוריים" in wb.sheetnames:
            ws_h = wb["תעריפי חשמל היסטוריים"]
            for r in range(3, ws_h.max_row + 1):
                d_val = ws_h.cell(row=r, column=1).value
                base_v = ws_h.cell(row=r, column=2).value
                vat_str = ws_h.cell(row=r, column=3).value
                agorot_val = ws_h.cell(row=r, column=5).value
                notes_val = ws_h.cell(row=r, column=6).value
                if d_val:
                    tariff_history.append({
                        "date": str(d_val),
                        "base_rate": base_v,
                        "vat": vat_str,
                        "rate_with_vat": round(float(base_v) * 1.17, 4) if isinstance(base_v, (int, float)) else 0,
                        "agorot": agorot_val,
                        "notes": notes_val
                    })

        return {
            "base_rate_g2": base_rate,
            "rate_with_vat_g1": rate_with_vat,
            "big_watch": big_watch_rows,
            "small_watch": small_watch_rows,
            "tariff_history": tariff_history
        }

    def update_base_tariff(self, base_rate):
        """Updates cell G2 in Excel with the base tariff before VAT."""
        self._create_backup()
        wb = openpyxl.load_workbook(self.file_path, data_only=False)
        sheet = wb.active
        sheet["G2"] = f"={base_rate}"
        # Ensure G1 is formula =G2*1.17
        sheet["G1"] = "=G2*1.17"
        wb.save(self.file_path)
        return True

    def add_meter_reading(self, meter_type, date_str, new_reading, photo_path=None, base_rate=None):
        """
        Adds a new meter reading row for either 'big_watch' (Tenant 1) or 'small_watch' (Tenant 2).
        Automatically formats Excel formulas and writes Column I or Column Q payment calculation.
        """
        self._create_backup()

        # Update G2 tariff if base_rate provided
        wb = openpyxl.load_workbook(self.file_path, data_only=False)
        sheet = wb.active

        if base_rate is not None:
            sheet["G2"] = float(base_rate)
            sheet["G1"] = "=G2*1.17"

        if meter_type in ["big_watch", "tenant_1", "main"]:
            # Main Meter / Big Watch
            # Find next empty row for Big Watch (checking Col C / Col D)
            r = 4
            while sheet.cell(row=r, column=4).value is not None or sheet.cell(row=r, column=3).value is not None:
                r += 1

            prev_row = r - 1
            # Column B: =B{prev}+1
            sheet.cell(row=r, column=2, value=f"=B{prev_row}+1")
            # Column C: Date
            sheet.cell(row=r, column=3, value=date_str)
            # Column D: Reading
            sheet.cell(row=r, column=4, value=float(new_reading))
            # Column E: kwh month =D{r}-D{prev}
            sheet.cell(row=r, column=5, value=f"=D{r}-D{prev_row}")
            # Column F: Days =DAYS(C{r}, C{prev})
            sheet.cell(row=r, column=6, value=f"=DAYS(C{r},C{prev_row})")
            # Column G: Kwh a day =E{r}/F{r}
            sheet.cell(row=r, column=7, value=f"=E{r}/F{r}")
            # Column H: Differences B-S =E{r}-N{r}
            sheet.cell(row=r, column=8, value=f"=E{r}-N{r}")
            # Column I: monthly payments =H{r}*$G$1
            sheet.cell(row=r, column=9, value=f"=H{r}*$G$1")

            result_target = f"Column I (Row {r})"

        else:
            # Sub Meter / Small Watch (Tenant 2)
            # Find next empty row for Small Watch (checking Col L / Col M)
            r = 4
            while sheet.cell(row=r, column=13).value is not None or sheet.cell(row=r, column=12).value is not None:
                r += 1

            prev_row = r - 1
            # Column K: =K{prev}+1
            sheet.cell(row=r, column=11, value=f"=K{prev_row}+1")
            # Column L: Date
            sheet.cell(row=r, column=12, value=date_str)
            # Column M: Reading
            sheet.cell(row=r, column=13, value=float(new_reading))
            # Column N: kwh month =M{r}-M{prev}
            sheet.cell(row=r, column=14, value=f"=M{r}-M{prev_row}")
            # Column O: Days =DAYS(L{r}, L{prev})
            sheet.cell(row=r, column=15, value=f"=DAYS(L{r},L{prev_row})")
            # Column P: Kwh a day =N{r}/O{r}
            sheet.cell(row=r, column=16, value=f"=N{r}/O{r}")
            # Column Q: monthly payments =N{r}*$G$1
            sheet.cell(row=r, column=17, value=f"=N{r}*$G$1")

            result_target = f"Column Q (Row {r})"

        wb.save(self.file_path)
        return {
            "success": True,
            "row": r,
            "meter_type": meter_type,
            "target": result_target,
            "date": date_str,
            "reading": float(new_reading),
            "photo_path": photo_path
        }

excel_handler = ExcelHandler()

if __name__ == "__main__":
    hist = excel_handler.get_history()
    print(f"Loaded {len(hist['big_watch'])} Big Watch rows, {len(hist['small_watch'])} Small Watch rows.")
